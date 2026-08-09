# pyrefly: ignore [missing-import]
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
# pyrefly: ignore [missing-import]
import numpy as np
import json
import os
import csv
import io
import openpyxl
import sqlite3

from engine.permutations import gerar_cases
from engine.normalization import escala_razao
from engine.surrogate import calcular_surrogate, extract_unique_solutions, calcular_pesos_roc
from engine.promethee_results import promethee_results
from engine.stats import compute_statistics
from engine.decision_rules import apply_decision_rules
from engine.elicitation import analise_para_elicitacao
from engine.sensitivity import run_sensitivity_analysis

app = Flask(__name__)
# In a real app, use a strong secret key. For local dev, a simple one is fine.
app.secret_key = 'spear_secret_key_local'

DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'spear.db')

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuario (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                validation TEXT DEFAULT 'validado'
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS problema (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_problema TEXT NOT NULL,
                data_problema TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                ID_usuario INTEGER NOT NULL,
                FOREIGN KEY(ID_usuario) REFERENCES usuario(id)
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS criterio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_criterio TEXT NOT NULL,
                tipo_criterio INTEGER NOT NULL,
                niveis INTEGER NOT NULL,
                ID_problema INTEGER NOT NULL,
                FOREIGN KEY(ID_problema) REFERENCES problema(id)
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alternativa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_alternativa TEXT NOT NULL,
                ID_problema INTEGER NOT NULL,
                FOREIGN KEY(ID_problema) REFERENCES problema(id)
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS matrizconsequencia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_alternativa INTEGER NOT NULL,
                ID_criterio INTEGER NOT NULL,
                valor_performance REAL NOT NULL,
                ID_problema INTEGER NOT NULL,
                FOREIGN KEY(ID_alternativa) REFERENCES alternativa(id),
                FOREIGN KEY(ID_criterio) REFERENCES criterio(id),
                FOREIGN KEY(ID_problema) REFERENCES problema(id)
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS avaliacaoholistica (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_problema INTEGER NOT NULL,
                alt1_nome TEXT NOT NULL,
                alt2_nome TEXT NOT NULL,
                tipo_relacao TEXT NOT NULL,
                FOREIGN KEY(ID_problema) REFERENCES problema(id)
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS decomposicaopreferencia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_problema INTEGER NOT NULL,
                criterio_a TEXT NOT NULL,
                criterio_b TEXT NOT NULL,
                tipo_relacao TEXT NOT NULL,
                valor_ratio REAL NOT NULL,
                FOREIGN KEY(ID_problema) REFERENCES problema(id)
            );
        ''')
        conn.commit()
        
        try:
            cursor.execute("ALTER TABLE problema ADD COLUMN racionalidade TEXT DEFAULT 'compensatory'")
            conn.commit()
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE avaliacaoholistica ADD COLUMN fictitious_value REAL DEFAULT NULL")
            conn.commit()
        except sqlite3.OperationalError:
            pass

init_db()

class NumpyEncoder(json.JSONEncoder):
    """ Custom encoder for numpy data types """
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)

app.json_encoder = NumpyEncoder

def search_next_decomposition_question(weights_matrix, active_valid_mask, num_crit, nomes_crit, matriz_poa, excluded_set):
    num_cases = weights_matrix.shape[0]
    num_alt = matriz_poa.shape[1]
    
    best_pair = None
    best_ratio = None
    best_score = -1.0
    best_split_diff = float('inf')
    
    # We only loop over normal permutations (excluding the last equal weights row)
    valid_perm_indices = [i for i in range(num_cases - 1) if active_valid_mask[i]]
    
    if len(valid_perm_indices) > 0:
        for a in range(num_crit):
            for b in range(num_crit):
                if a == b:
                    continue
                if (a, b) in excluded_set or (b, a) in excluded_set:
                    continue
                    
                # Get weights for criterion a and b for valid permutations
                w_a = weights_matrix[valid_perm_indices, a]
                w_b = weights_matrix[valid_perm_indices, b]
                
                # Filter indices where w_a > 0 and w_b <= w_a to avoid division by zero and ensure Y <= 1
                valid_ratio_sub_indices = [idx for idx, val in enumerate(w_a) if val > 0 and w_b[idx] <= val + 1e-9]
                if not valid_ratio_sub_indices:
                    continue
                    
                ratios = w_b[valid_ratio_sub_indices] / w_a[valid_ratio_sub_indices]
                
                # Get unique ratios (with tolerance)
                unique_ratios = []
                for r in ratios:
                    if not any(np.isclose(r, ur, atol=1e-9) for ur in unique_ratios):
                        unique_ratios.append(r)
                
                for Y in unique_ratios:
                    Y = float(Y)
                    if Y > 1.0 + 1e-9:
                        continue
                    
                    # Outcome 1: w_b > Y * w_a (strict inequality)
                    mask1 = w_b > Y * w_a + 1e-9
                    indices1 = [valid_perm_indices[idx] for idx, val in enumerate(mask1) if val]
                    
                    # Outcome 2: w_b <= Y * w_a (non-strict inequality)
                    mask2 = w_b <= Y * w_a + 1e-9
                    indices2 = [valid_perm_indices[idx] for idx, val in enumerate(mask2) if val]
                    
                    if not indices1 or not indices2:
                        continue
                        
                    # Compute max probability for Outcome 1
                    poa_subset1 = matriz_poa[indices1, :]
                    probs1 = np.sum(poa_subset1, axis=0) / len(indices1)
                    p1 = float(np.max(probs1))
                    
                    # Compute max probability for Outcome 2
                    poa_subset2 = matriz_poa[indices2, :]
                    probs2 = np.sum(poa_subset2, axis=0) / len(indices2)
                    p2 = float(np.max(probs2))
                    
                    score = min(p1, p2)
                    split_diff = abs(len(indices1) - len(indices2))
                    
                    if score > best_score:
                        best_score = score
                        best_pair = (a, b)
                        best_ratio = Y
                        best_split_diff = split_diff
                    elif np.isclose(score, best_score, atol=1e-9):
                        if split_diff < best_split_diff:
                            best_pair = (a, b)
                            best_ratio = Y
                            best_split_diff = split_diff
                        
    if best_pair is not None:
        return {
            'critA': nomes_crit[best_pair[0]],
            'critB': nomes_crit[best_pair[1]],
            'critAIdx': best_pair[0],
            'critBIdx': best_pair[1],
            'ratio': best_ratio
        }
    return None

@app.route('/')
def index():
    return render_template('welcome.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        next_url = request.args.get('next') or url_for('options')
        return redirect(next_url)
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        next_url = request.form.get('next') or url_for('options')
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM usuario WHERE email = ? AND password = ?", (email, password))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            session['user_id'] = user['id']
            session['user_email'] = user['email']
            return redirect(next_url)
        else:
            return render_template('login.html', error="Invalid email or password.", next=next_url)
            
    return render_template('login.html', next=request.args.get('next', ''))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not email or not password:
            return render_template('register.html', error="Email and password are required.")
        if password != confirm_password:
            return render_template('register.html', error="Passwords do not match.")
            
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO usuario (email, password, validation) VALUES (?, ?, 'validado')", (email, password))
            conn.commit()
            conn.close()
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            return render_template('register.html', error="Email already exists.")
            
    return render_template('register.html')

@app.route('/options')
def options():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('options.html')

@app.route('/continue')
def continue_problem():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome_problema, data_problema FROM problema WHERE ID_usuario = ? ORDER BY data_problema DESC", (session['user_id'],))
    problems = cursor.fetchall()
    conn.close()
    
    return render_template('continue.html', problems=problems)

@app.route('/setup')
def setup():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('setup.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/results')
def results():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('results.html')

@app.route('/api/import', methods=['POST'])
def import_data():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    
    file = request.files['file']
    filename = file.filename.lower()
    
    try:
        # Initialize containers
        criteria = []
        crit_types = []
        niveis = []
        alternatives = []
        matrix = []
        
        # CSV handling (FITradeoff format)
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode('utf-8-sig'), newline=None)
            sample = stream.read(1024)
            stream.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';'])
            reader = csv.reader(stream, dialect)
            rows = list(reader)
            # Row 1: criteria names from column B onward
            criteria = [c.strip() for c in rows[0][1:] if c.strip()]
            # Row 2: criterion types
            criterion_types = [int(t.strip()) for t in rows[1][1:1+len(criteria)] if t.strip()]
            # Row 7 (index 6): levels for discrete criteria (only types 2 or 3)
            levels = []
            if len(rows) > 6:
                for idx, ct in enumerate(criterion_types):
                    if ct in [2, 3]:
                        val = rows[6][idx+1] if len(rows[6]) > idx+1 else ''
                        try:
                            levels.append(int(val))
                        except:
                            levels.append(0)
                    else:
                        levels.append(0)
            else:
                levels = [0] * len(criteria)
            # Alternatives start at row 9 (index 8) column A
            alternatives = []
            matrix = []
            alt_start = 8
            while alt_start < len(rows) and rows[alt_start][0]:
                alternatives.append(str(rows[alt_start][0]).strip())
                # Matrix row values start at column B
                row_vals = []
                for val in rows[alt_start][1:1+len(criteria)]:
                    try:
                        row_vals.append(float(val))
                    except:
                        row_vals.append(0.0)
                matrix.append(row_vals)
                alt_start += 1
            return jsonify({
                'success': True,
                'criteria': criteria,
                'criterionTypes': criterion_types,
                'levels': levels,
                'alternatives': alternatives,
                'matrix': matrix
            })
        # XLSX handling (FITradeoff format)
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active
            # Determine number of criteria by scanning row 1 starting column B
            criteria = []
            col = 2
            while True:
                val = sheet.cell(row=1, column=col).value
                if val is None:
                    break
                criteria.append(str(val).strip())
                col += 1
            # Row 2: criterion types
            criterion_types = []
            for idx in range(len(criteria)):
                val = sheet.cell(row=2, column=2+idx).value
                try:
                    criterion_types.append(int(val))
                except:
                    criterion_types.append(0)
            # Row 7: levels
            levels = []
            for idx in range(len(criteria)):
                val = sheet.cell(row=7, column=2+idx).value
                try:
                    levels.append(int(val))
                except:
                    levels.append(0)
            # Alternatives start at row 9 column A
            alternatives = []
            matrix = []
            row_idx = 9
            while True:
                alt_val = sheet.cell(row=row_idx, column=1).value
                if alt_val is None:
                    break
                alternatives.append(str(alt_val).strip())
                # Matrix row values start at column B
                row_vals = []
                for idx in range(len(criteria)):
                    cell = sheet.cell(row=row_idx, column=2+idx)
                    try:
                        row_vals.append(float(cell.value))
                    except:
                        row_vals.append(0.0)
                matrix.append(row_vals)
                row_idx += 1
            return jsonify({
                'success': True,
                'criteria': criteria,
                'criterionTypes': criterion_types,
                'levels': levels,
                'alternatives': alternatives,
                'matrix': matrix
            })
        else:
            return jsonify({'success': False, 'error': 'Unsupported file format.'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/solve', methods=['POST'])
def solve():
    data = request.json

    try:
        problem_name = data.get('problemName', 'SPEAR Problem')
        num_crit = int(data['numCrit'])
        num_alt = int(data['numAlt'])

        # Criteria metadata
        tipocrit = [int(x) for x in data['tipoCrit']]
        niveis = [int(x) for x in data.get('niveisCrit', [0]*num_crit)]
        nomes_crit = data.get('nomeCrit', [f'C{i+1}' for i in range(num_crit)])
        nomes_alt = data.get('nomeAlt', [f'A{i+1}' for i in range(num_alt)])

        # Apply stored criteria ordering if present
        order = session.get('criteria_order')
        if order and len(order) == num_crit:
            tipocrit = [tipocrit[i] for i in order]
            niveis = [niveis[i] for i in order]
            nomes_crit = [nomes_crit[i] for i in order]

        # Consequence matrix
        matriz_conseq = np.array(data['matrizConseq'], dtype=float)
        
        # 1. Permutations
        cases_ordem_crit = gerar_cases(num_crit)
        
        # Apply rank position filters if present
        rank_filters = data.get('rankFilters')
        if rank_filters:
            perms = cases_ordem_crit[:-1]
            valid_mask = np.ones(perms.shape[0], dtype=bool)
            for p, crit_idx in enumerate(rank_filters):
                if crit_idx is not None:
                    crit_idx = int(crit_idx)
                    valid_mask &= (perms[:, crit_idx] == (p + 1))
            
            filtered_perms = perms[valid_mask]
            if filtered_perms.shape[0] == 0:
                return jsonify({
                    'success': False,
                    'error': 'Inconsistency detected: The criteria rank filters leave 0 valid permutations.'
                })
            cases_ordem_crit = np.vstack([filtered_perms, cases_ordem_crit[-1]])
        
        # 2. Normalization
        matriz_conseq_norm, max_crit, min_crit = escala_razao(matriz_conseq, tipocrit, niveis)
        
        # 3. Surrogate Engine (ROC & Promethee evaluations)
        surrogate_data = calcular_surrogate(cases_ordem_crit, matriz_conseq_norm, tipocrit)
        
        # ── Model Selection ────────────────────────────────────────────────────
        # The two models (additive/ROC and outranking/PROMETHEE) are completely
        # independent. Selecting one rationality means ONLY that model is used.
        # The other model is not computed, not filtered, and not returned.
        rationality = data.get('rationality', 'compensatory')
        is_compensatory = (rationality == 'compensatory')

        # ── Holistic Evaluation Filtering (active model only) ──────────────────
        holistic_evals = data.get('holisticEvaluations', [])
        num_cases = cases_ordem_crit.shape[0]
        active_valid_mask = np.ones(num_cases, dtype=bool)

        nomes_alt_clean = [str(x).strip() for x in nomes_alt]

        for ev in holistic_evals:
            alt1_name = str(ev.get('alt1', '')).strip()
            alt2_name = str(ev.get('alt2', '')).strip()
            relation   = ev.get('relation', '')

            if not alt1_name or not alt2_name or relation not in ['>=', '<=']:
                continue

            idx1 = next((i for i, n in enumerate(nomes_alt_clean) if n == alt1_name), None)
            if alt2_name == 'fictitious':
                idx2 = 'fictitious'
            else:
                idx2 = next((i for i, n in enumerate(nomes_alt_clean) if n == alt2_name), None)

            if idx1 is None or idx2 is None:
                continue

            if is_compensatory:
                scores = surrogate_data['resultado_roc']
            else:
                scores = surrogate_data['resultado_promethee']

            if idx2 == 'fictitious':
                fict_val = ev.get('fictitiousValue')
                if fict_val is None:
                    # Fallback: compute minimum global value of a solution across all cases
                    fict_val = float(np.min(np.max(scores, axis=1)))
                else:
                    fict_val = float(fict_val)
                    
                if relation == '>=':
                    active_valid_mask &= (scores[:, idx1] >= fict_val - 1e-9)
                else:
                    active_valid_mask &= (scores[:, idx1] <= fict_val + 1e-9)
            else:
                if relation == '>=':
                    active_valid_mask &= (scores[:, idx1] >= scores[:, idx2] - 1e-9)
                else:
                    active_valid_mask &= (scores[:, idx1] <= scores[:, idx2] + 1e-9)

        # ── Decomposition Evaluation Filtering (active model only) ─────────────
        roc_w = calcular_pesos_roc(num_crit)
        weights_matrix = np.zeros((num_cases, num_crit))
        for k in range(num_cases - 1):
            for i in range(num_crit):
                weights_matrix[k, i] = roc_w[cases_ordem_crit[k, i] - 1]
        weights_matrix[-1, :] = 1.0 / num_crit

        decomposition_prefs = data.get('decompositionPreferences', [])
        for dp in decomposition_prefs:
            crit_a_name = dp.get('critA')
            crit_b_name = dp.get('critB')
            relation = dp.get('relation')
            ratio = float(dp.get('ratio', 0.0))
            
            idx_a = next((i for i, n in enumerate(nomes_crit) if n == crit_a_name), None)
            idx_b = next((i for i, n in enumerate(nomes_crit) if n == crit_b_name), None)
            
            if idx_a is not None and idx_b is not None and relation in ['>=', '<=']:
                w_a = weights_matrix[:, idx_a]
                w_b = weights_matrix[:, idx_b]
                if relation == '>=':
                    active_valid_mask &= (w_b >= ratio * w_a - 1e-9)
                else:
                    active_valid_mask &= (w_b <= ratio * w_a + 1e-9)

        # Validate: active model must have at least one valid permutation
        active_total_permuted = int(np.sum(active_valid_mask[:-1])) if num_cases > 1 else 0
        if active_total_permuted == 0:
            model_name = 'Modelo Aditivo (ROC)' if is_compensatory else 'Sobreclassificação (PROMETHEE)'
            return jsonify({
                'success': False,
                'error': (f'Inconsistência detectada: As preferências holísticas ou de decomposição são '
                          f'incompatíveis com os dados do {model_name}, não restando '
                          f'nenhum caso de pesos válido. Revise as preferências.')
            })

        # ── Apply filter and compute ONLY the active model ─────────────────────
        cases_filtered  = cases_ordem_crit[active_valid_mask]

        if is_compensatory:
            # ── Additive / ROC model ───────────────────────────────────────────
            resultado_active    = surrogate_data['resultado_roc'][active_valid_mask]
            matriz_dif_vg_act   = surrogate_data['matriz_dif_vg'][active_valid_mask]
            matriz_poa_active   = surrogate_data['matriz_poa'][active_valid_mask]

            matriz_sol, result_sol, case_sol = extract_unique_solutions(matriz_poa_active)
            stats     = compute_statistics(matriz_sol, result_sol, case_sol, matriz_dif_vg_act)
            total_cases = int(np.sum(result_sol))
            dec_rules = apply_decision_rules(result_sol, matriz_sol, stats, total_cases)

            elicitation_data = analise_para_elicitacao(
                cases_filtered, matriz_poa_active, result_sol, matriz_sol
            )

            # Empty placeholder for the inactive model
            empty_sol  = np.zeros((1, num_alt), dtype=int)
            empty_rsol = np.zeros(1)
            empty_poa  = np.zeros((1, num_alt), dtype=int)
            empty_res  = np.zeros((1, num_alt))

            roc_payload = {
                'totalCases': total_cases,
                'matrizSol':  matriz_sol.tolist(),
                'resultSol':  result_sol.tolist(),
                'stats':      {k: v.tolist() if isinstance(v, np.ndarray) else v for k, v in stats.items()},
                'decisionRule': dec_rules
            }
            promethee_payload = {
                'totalCases': 0,
                'matrizSol':  empty_sol.tolist(),
                'resultSol':  empty_rsol.tolist(),
                'stats':      {},
                'decisionRule': {'status': 'N/A', 'recommended_alts': [], 'probability': 0}
            }
            raw_poa_roc        = matriz_poa_active.tolist()
            raw_poa_promethee  = empty_poa.tolist()
            raw_resultado_roc  = resultado_active.tolist()
            raw_resultado_prom = empty_res.tolist()
            cases_roc_raw      = cases_filtered.tolist()
            cases_prom_raw     = []

        else:
            # ── Outranking / PROMETHEE model ───────────────────────────────────
            resultado_active    = surrogate_data['resultado_promethee'][active_valid_mask]
            matriz_dif_vg_act   = surrogate_data['matriz_dif_vg_promethee'][active_valid_mask]
            matriz_poa_active   = surrogate_data['matriz_poa_promethee'][active_valid_mask]

            matriz_sol, result_sol, case_sol = extract_unique_solutions(matriz_poa_active)
            stats     = compute_statistics(matriz_sol, result_sol, case_sol, matriz_dif_vg_act)
            total_cases = int(np.sum(result_sol))
            dec_rules = apply_decision_rules(result_sol, matriz_sol, stats, total_cases)

            # Elicitation is ROC-based (weight ordering concept); still useful
            # even in non-compensatory mode as a reference, but uses unfiltered
            # ROC permutations independently — no cross-model contamination.
            elicitation_data = analise_para_elicitacao(
                cases_ordem_crit,
                surrogate_data['matriz_poa'],
                np.zeros(len(cases_ordem_crit), dtype=int),  # neutral placeholder
                np.zeros((1, num_alt), dtype=int)
            )

            empty_sol  = np.zeros((1, num_alt), dtype=int)
            empty_rsol = np.zeros(1)
            empty_poa  = np.zeros((1, num_alt), dtype=int)
            empty_res  = np.zeros((1, num_alt))

            promethee_payload = {
                'totalCases': total_cases,
                'matrizSol':  matriz_sol.tolist(),
                'resultSol':  result_sol.tolist(),
                'stats':      {k: v.tolist() if isinstance(v, np.ndarray) else v for k, v in stats.items()},
                'decisionRule': dec_rules
            }
            roc_payload = {
                'totalCases': 0,
                'matrizSol':  empty_sol.tolist(),
                'resultSol':  empty_rsol.tolist(),
                'stats':      {},
                'decisionRule': {'status': 'N/A', 'recommended_alts': [], 'probability': 0}
            }
            raw_poa_roc        = empty_poa.tolist()
            raw_poa_promethee  = matriz_poa_active.tolist()
            raw_resultado_roc  = empty_res.tolist()
            raw_resultado_prom = resultado_active.tolist()
            cases_roc_raw      = []
            cases_prom_raw     = cases_filtered.tolist()

        # Compute next decomposition question if not converged
        converged = (len(matriz_sol) == 1)
        excluded_pairs = data.get('excludedPairs', [])
        excluded_set = {(p[0], p[1]) for p in excluded_pairs}
        
        next_q = None
        if not converged:
            next_q = search_next_decomposition_question(
                weights_matrix,
                active_valid_mask,
                num_crit,
                nomes_crit,
                surrogate_data['matriz_poa'] if is_compensatory else surrogate_data['matriz_poa_promethee'],
                excluded_set
            )

        # ── Build JSON response ────────────────────────────────────────────────
        response = {
            'success': True,
            'problemName': problem_name,
            'nomesAlt':    nomes_alt,
            'nomesCrit':   nomes_crit,
            'totalCases':  total_cases,
            'matrizConseq':     matriz_conseq.tolist(),
            'matrizConseqNorm': matriz_conseq_norm.tolist(),
            'tipoCrit':    tipocrit,
            'niveisCrit':  niveis,
            'rationality': rationality,
            'holisticEvaluations': holistic_evals,
            'decompositionQuestion': next_q,

            'roc':       roc_payload,
            'promethee': promethee_payload,

            'elicitation': {
                'altX':           elicitation_data['alt_x'],
                'altZ':           elicitation_data['alt_z'],
                'matrizProbX':    elicitation_data['matriz_prob_x'].tolist(),
                'matrizProbZ':    elicitation_data['matriz_prob_z'].tolist(),
                'matrizProbOutros': elicitation_data['matriz_prob_outros'].tolist()
            },
            'raw': {
                'casesOrdemCritRoc':      cases_roc_raw,
                'casesOrdemCritPromethee': cases_prom_raw,
                'resultadoRoc':       raw_resultado_roc,
                'resultadoPromethee': raw_resultado_prom,
                'matrizPoa':          raw_poa_roc,
                'matrizPoaPromethee': raw_poa_promethee
            }
        }

        return jsonify(response)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/sensitivity', methods=['POST'])
def sensitivity():
    data = request.json
    try:
        num_crit = int(data['numCrit'])
        num_alt = int(data['numAlt'])
        
        # Criteria metadata
        tipocrit = [int(x) for x in data['tipoCrit']]
        niveis = [int(x) for x in data.get('niveisCrit', [0]*num_crit)]
        nomes_crit = data.get('nomeCrit', [f'C{i+1}' for i in range(num_crit)])
        nomes_alt = data.get('nomeAlt', [f'A{i+1}' for i in range(num_alt)])
        
        # Consequence matrix
        matriz_conseq = np.array(data['matrizConseq'], dtype=float)
        
        # Active filters
        rank_filters = data.get('rankFilters')
        holistic_evals = data.get('holisticEvaluations', [])
        decomp_prefs = data.get('decompositionPreferences', [])
        
        # Variations (percentages)
        variations_pct = [float(v) for v in data['variationsPct']]
        
        rationality = data.get('rationality', 'compensatory')
        
        # Call sensitivity analysis engine
        res = run_sensitivity_analysis(
            matriz_conseq=matriz_conseq,
            tipocrit=tipocrit,
            niveis=niveis,
            rationality=rationality,
            rank_filters=rank_filters,
            holistic_evals=holistic_evals,
            nomes_alt=nomes_alt,
            variations_pct=variations_pct,
            num_simulations=10000,
            nomes_crit=nomes_crit,
            decomposition_prefs=decomp_prefs
        )
        
        return jsonify({
            'success': True,
            'alternatives': res['alternatives'],
            'probabilities': res['probabilities'],
            'deltas': res['deltas'],
            'min_crit': res['min_crit'],
            'max_crit': res['max_crit']
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/save_problem', methods=['POST'])
def save_problem():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    data = request.json
    try:
        problem_id = data.get('problemId') # If editing existing
        problem_name = data.get('problemName', 'SPEAR Problem')
        rationality = data.get('rationality', 'compensatory')
        criteria = data.get('criteria', [])
        criterion_types = data.get('criterionTypes', [])
        levels = data.get('levels', [])
        alternatives = data.get('alternatives', [])
        matrix = data.get('matrix', [])
        
        conn = get_db()
        cursor = conn.cursor()
        
        if problem_id:
            # Check ownership
            cursor.execute("SELECT id FROM problema WHERE id = ? AND ID_usuario = ?", (problem_id, session['user_id']))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'error': 'Unauthorized'}), 401
            # Update problem name and rationality
            cursor.execute("UPDATE problema SET nome_problema = ?, racionalidade = ? WHERE id = ?", (problem_name, rationality, problem_id))
            # Delete old criteria, alternatives, matrix elements
            cursor.execute("DELETE FROM matrizconsequencia WHERE ID_problema = ?", (problem_id,))
            cursor.execute("DELETE FROM criterio WHERE ID_problema = ?", (problem_id,))
            cursor.execute("DELETE FROM alternativa WHERE ID_problema = ?", (problem_id,))
        else:
            # Create new problem record
            cursor.execute("INSERT INTO problema (nome_problema, ID_usuario, racionalidade) VALUES (?, ?, ?)", (problem_name, session['user_id'], rationality))
            problem_id = cursor.lastrowid
            
        # Insert criteria
        crit_ids = []
        for i, name in enumerate(criteria):
            t_crit = int(criterion_types[i]) if i < len(criterion_types) else 0
            niveis_val = int(levels[i]) if i < len(levels) else 0
            cursor.execute(
                "INSERT INTO criterio (nome_criterio, tipo_criterio, niveis, ID_problema) VALUES (?, ?, ?, ?)",
                (name, t_crit, niveis_val, problem_id)
            )
            crit_ids.append(cursor.lastrowid)
            
        # Insert alternatives
        alt_ids = []
        for name in alternatives:
            cursor.execute(
                "INSERT INTO alternativa (nome_alternativa, ID_problema) VALUES (?, ?)",
                (name, problem_id)
            )
            alt_ids.append(cursor.lastrowid)
            
        # Insert consequence matrix values
        for alt_idx, row in enumerate(matrix):
            for crit_idx, val in enumerate(row):
                if alt_idx < len(alt_ids) and crit_idx < len(crit_ids):
                    cursor.execute(
                        "INSERT INTO matrizconsequencia (ID_alternativa, ID_criterio, valor_performance, ID_problema) VALUES (?, ?, ?, ?)",
                        (alt_ids[alt_idx], crit_ids[crit_idx], float(val), problem_id)
                    )
                    
        # Delete old holistic evaluations if editing
        cursor.execute("DELETE FROM avaliacaoholistica WHERE ID_problema = ?", (problem_id,))
        # Insert new holistic evaluations
        holistic_evals = data.get('holisticEvaluations', [])
        for ev in holistic_evals:
            alt1 = ev.get('alt1')
            alt2 = ev.get('alt2')
            relation = ev.get('relation')
            fict_val = ev.get('fictitiousValue')
            if alt1 and alt2 and relation in ['>=', '<=']:
                cursor.execute(
                    "INSERT INTO avaliacaoholistica (ID_problema, alt1_nome, alt2_nome, tipo_relacao, fictitious_value) VALUES (?, ?, ?, ?, ?)",
                    (problem_id, alt1, alt2, relation, fict_val)
                )
                    
        # Delete old decomposition preferences if editing
        cursor.execute("DELETE FROM decomposicaopreferencia WHERE ID_problema = ?", (problem_id,))
        # Insert new decomposition preferences
        decomp_prefs = data.get('decompositionPreferences', [])
        for dp in decomp_prefs:
            crit_a = dp.get('critA')
            crit_b = dp.get('critB')
            relation = dp.get('relation')
            ratio = dp.get('ratio')
            if crit_a and crit_b and relation in ['>=', '<='] and ratio is not None:
                cursor.execute(
                    "INSERT INTO decomposicaopreferencia (ID_problema, criterio_a, criterio_b, tipo_relacao, valor_ratio) VALUES (?, ?, ?, ?, ?)",
                    (problem_id, crit_a, crit_b, relation, float(ratio))
                )

        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'problemId': problem_id})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/load_problem/<int:id>', methods=['GET'])
def load_problem(id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    # Ensure problem belongs to the logged in user
    cursor.execute("SELECT * FROM problema WHERE id = ? AND ID_usuario = ?", (id, session['user_id']))
    problem = cursor.fetchone()
    if not problem:
        conn.close()
        return jsonify({'success': False, 'error': 'Problem not found'}), 404
        
    # Get criteria
    cursor.execute("SELECT id, nome_criterio, tipo_criterio, niveis FROM criterio WHERE ID_problema = ? ORDER BY id", (id,))
    criteria_rows = cursor.fetchall()
    
    # Get alternatives
    cursor.execute("SELECT id, nome_alternativa FROM alternativa WHERE ID_problema = ? ORDER BY id", (id,))
    alt_rows = cursor.fetchall()
    
    # Create mappings of IDs to list indexes
    crit_id_map = {row['id']: idx for idx, row in enumerate(criteria_rows)}
    alt_id_map = {row['id']: idx for idx, row in enumerate(alt_rows)}
    
    num_crit = len(criteria_rows)
    num_alt = len(alt_rows)
    
    # Get consequence matrix
    cursor.execute("SELECT ID_alternativa, ID_criterio, valor_performance FROM matrizconsequencia WHERE ID_problema = ?", (id,))
    matrix_rows = cursor.fetchall()
    
    # Get holistic evaluations
    cursor.execute("SELECT alt1_nome, alt2_nome, tipo_relacao, fictitious_value FROM avaliacaoholistica WHERE ID_problema = ? ORDER BY id", (id,))
    holistic_rows = cursor.fetchall()
    
    # Get decomposition preferences
    cursor.execute("SELECT criterio_a, criterio_b, tipo_relacao, valor_ratio FROM decomposicaopreferencia WHERE ID_problema = ? ORDER BY id", (id,))
    decomp_rows = cursor.fetchall()
    
    conn.close()
    
    # Reconstruct matrix
    matrix = [[0.0 for _ in range(num_crit)] for _ in range(num_alt)]
    for row in matrix_rows:
        alt_idx = alt_id_map.get(row['ID_alternativa'])
        crit_idx = crit_id_map.get(row['ID_criterio'])
        if alt_idx is not None and crit_idx is not None:
            matrix[alt_idx][crit_idx] = row['valor_performance']
            
    return jsonify({
        'success': True,
        'problemId': problem['id'],
        'problemName': problem['nome_problema'],
        'rationality': problem['racionalidade'] or 'compensatory',
        'criteria': [r['nome_criterio'] for r in criteria_rows],
        'criterionTypes': [r['tipo_criterio'] for r in criteria_rows],
        'levels': [r['niveis'] for r in criteria_rows],
        'alternatives': [r['nome_alternativa'] for r in alt_rows],
        'matrix': matrix,
        'holisticEvaluations': [{'alt1': r['alt1_nome'], 'alt2': r['alt2_nome'], 'relation': r['tipo_relacao'], 'fictitiousValue': r['fictitious_value']} for r in holistic_rows],
        'decompositionPreferences': [{'critA': r['criterio_a'], 'critB': r['criterio_b'], 'relation': r['tipo_relacao'], 'ratio': r['valor_ratio']} for r in decomp_rows]
    })

@app.route('/api/log', methods=['POST'])
def client_log():
    data = request.json
    message = data.get('message', '')
    print(f"[CLIENT LOG]: {message}", flush=True)
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True)
